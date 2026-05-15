from __future__ import annotations

import json
import re
from dataclasses import asdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from legal_anonymizer.models import Entity, MappingTable, PlaceholderMapping, utc_timestamp


def build_mapping_table(source_name: str, entities: list[Entity]) -> MappingTable:
    counters: dict[str, int] = {}
    by_key: dict[tuple[str, str], PlaceholderMapping] = {}
    mappings: list[PlaceholderMapping] = []

    for entity in sorted(entities, key=lambda item: (item.category, item.value)):
        key = (entity.category, entity.value)
        if key in by_key:
            continue
        counters[entity.category] = counters.get(entity.category, 0) + 1
        placeholder = f"[[{entity.category}_{counters[entity.category]:03d}]]"
        sources = sorted({item.source for item in entities if item.category == entity.category and item.value == entity.value})
        mapping = PlaceholderMapping(
            placeholder=placeholder,
            category=entity.category,
            value=entity.value,
            sources=sources,
        )
        by_key[key] = mapping
        mappings.append(mapping)

    return MappingTable(source_name=source_name, created_at=utc_timestamp(), mappings=mappings)


def save_mapping_table(mapping_dir: Path, table: MappingTable) -> Path:
    mapping_dir.mkdir(parents=True, exist_ok=True)
    safe_stem = _safe_stem(Path(table.source_name).stem)
    path = mapping_dir / f"{safe_stem}.mapping.json"
    payload = asdict(table)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def export_mapping_xlsx(mapping_dir: Path, table: MappingTable) -> Path:
    mapping_dir.mkdir(parents=True, exist_ok=True)
    safe_stem = _safe_stem(Path(table.source_name).stem)
    path = mapping_dir / f"{safe_stem}.脱敏信息对照表.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "脱敏信息对照表"
    sheet.append(["", "原始内容", "脱敏后内容", "类型", "来源", "备注"])
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for mapping in table.mappings:
        sheet.append(["", mapping.value, mapping.placeholder, mapping.category, ", ".join(mapping.sources), ""])
    widths = {2: 36, 3: 24, 4: 16, 5: 28, 6: 24}
    for column_index, width in widths.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    workbook.save(path)
    return path


def load_mapping_table(path: Path) -> MappingTable:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    return MappingTable(
        source_name=payload["source_name"],
        created_at=payload["created_at"],
        mappings=[PlaceholderMapping(**item) for item in payload["mappings"]],
    )


def _safe_stem(stem: str) -> str:
    cleaned = re.sub(r"[^\w\u4e00-\u9fa5.-]+", "_", stem, flags=re.UNICODE).strip("._")
    return cleaned or "document"
